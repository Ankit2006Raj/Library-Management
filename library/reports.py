"""
Advanced Reporting System
Generate comprehensive reports in various formats
"""
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Book, BorrowRecord, User, UserProfile, Review


class ReportGenerator:
    """Generate various types of reports"""
    
    @staticmethod
    def generate_circulation_report(start_date, end_date):
        """Generate circulation report for a date range"""
        borrows = BorrowRecord.objects.filter(
            borrow_date__date__gte=start_date,
            borrow_date__date__lte=end_date
        )
        
        return {
            'period': f"{start_date} to {end_date}",
            'total_borrows': borrows.count(),
            'total_returns': borrows.filter(status='Returned').count(),
            'currently_borrowed': borrows.filter(status='Borrowed').count(),
            'overdue': borrows.filter(status='Overdue').count(),
            'total_fines': borrows.aggregate(Sum('fine_amount'))['fine_amount__sum'] or 0,
            'fines_collected': borrows.filter(fine_paid=True).aggregate(
                Sum('fine_amount')
            )['fine_amount__sum'] or 0,
            'unique_borrowers': borrows.values('user').distinct().count(),
            'unique_books': borrows.values('book').distinct().count(),
            'by_category': borrows.values('book__category').annotate(
                count=Count('id')
            ).order_by('-count'),
        }
    
    @staticmethod
    def generate_inventory_report():
        """Generate complete inventory report"""
        books = Book.objects.all()
        
        return {
            'total_books': books.count(),
            'total_copies': books.aggregate(Sum('total_copies'))['total_copies__sum'] or 0,
            'available_copies': books.aggregate(Sum('copies_available'))['copies_available__sum'] or 0,
            'by_status': books.values('status').annotate(count=Count('id')),
            'by_category': books.values('category').annotate(
                count=Count('id'),
                total_copies=Sum('total_copies')
            ).order_by('-count'),
            'by_language': books.values('language').annotate(count=Count('id')),
            'low_stock': books.filter(copies_available__lte=1, status='Available'),
            'maintenance_needed': books.filter(status='Maintenance'),
            'average_rating': books.aggregate(Avg('rating'))['rating__avg'] or 0,
        }
    
    @staticmethod
    def generate_user_activity_report(start_date, end_date):
        """Generate user activity report"""
        users = User.objects.filter(is_active=True)
        borrows = BorrowRecord.objects.filter(
            borrow_date__date__gte=start_date,
            borrow_date__date__lte=end_date
        )
        
        return {
            'period': f"{start_date} to {end_date}",
            'total_users': users.count(),
            'active_borrowers': borrows.values('user').distinct().count(),
            'new_users': users.filter(
                date_joined__date__gte=start_date,
                date_joined__date__lte=end_date
            ).count(),
            'by_membership': UserProfile.objects.values('membership_type').annotate(
                count=Count('id')
            ),
            'top_borrowers': User.objects.annotate(
                borrow_count=Count('borrowings', filter=Q(
                    borrowings__borrow_date__date__gte=start_date,
                    borrowings__borrow_date__date__lte=end_date
                ))
            ).order_by('-borrow_count')[:10],
            'users_with_fines': UserProfile.objects.filter(total_fines__gt=0).count(),
            'total_fines_outstanding': UserProfile.objects.aggregate(
                Sum('total_fines')
            )['total_fines__sum'] or 0,
        }
    
    @staticmethod
    def generate_financial_report(start_date, end_date):
        """Generate financial report"""
        borrows = BorrowRecord.objects.filter(
            borrow_date__date__gte=start_date,
            borrow_date__date__lte=end_date
        )
        
        return {
            'period': f"{start_date} to {end_date}",
            'total_fines_generated': borrows.aggregate(
                Sum('fine_amount')
            )['fine_amount__sum'] or 0,
            'fines_collected': borrows.filter(fine_paid=True).aggregate(
                Sum('fine_amount')
            )['fine_amount__sum'] or 0,
            'fines_pending': borrows.filter(fine_paid=False).aggregate(
                Sum('fine_amount')
            )['fine_amount__sum'] or 0,
            'transactions_count': borrows.filter(fine_amount__gt=0).count(),
            'average_fine': borrows.filter(fine_amount__gt=0).aggregate(
                Avg('fine_amount')
            )['fine_amount__avg'] or 0,
        }
    
    @staticmethod
    def generate_popular_books_report(limit=50):
        """Generate report of most popular books"""
        books = Book.objects.annotate(
            borrow_count=Count('borrow_history'),
            avg_rating=Avg('reviews__rating'),
            review_count=Count('reviews')
        ).order_by('-borrow_count')[:limit]
        
        return [{
            'title': book.title,
            'author': book.author,
            'category': book.category,
            'borrow_count': book.borrow_count,
            'avg_rating': round(book.avg_rating or 0, 2),
            'review_count': book.review_count,
            'status': book.status,
        } for book in books]
    
    @staticmethod
    def export_circulation_report_pdf(start_date, end_date):
        """Export circulation report as PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph('Circulation Report', title_style))
        elements.append(Paragraph(f'Period: {start_date} to {end_date}', styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Get data
        report_data = ReportGenerator.generate_circulation_report(start_date, end_date)
        
        # Summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Borrows', str(report_data['total_borrows'])],
            ['Total Returns', str(report_data['total_returns'])],
            ['Currently Borrowed', str(report_data['currently_borrowed'])],
            ['Overdue Books', str(report_data['overdue'])],
            ['Total Fines', f"${report_data['total_fines']:.2f}"],
            ['Fines Collected', f"${report_data['fines_collected']:.2f}"],
            ['Unique Borrowers', str(report_data['unique_borrowers'])],
            ['Unique Books', str(report_data['unique_books'])],
        ]
        
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def export_inventory_report_excel():
        """Export inventory report as Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Header styling
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Headers
        headers = ['Title', 'Author', 'Category', 'ISBN', 'Total Copies', 
                  'Available', 'Status', 'Rating', 'Language']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        books = Book.objects.all().order_by('category', 'title')
        for row, book in enumerate(books, 2):
            ws.cell(row=row, column=1, value=book.title)
            ws.cell(row=row, column=2, value=book.author)
            ws.cell(row=row, column=3, value=book.category)
            ws.cell(row=row, column=4, value=book.isbn or 'N/A')
            ws.cell(row=row, column=5, value=book.total_copies)
            ws.cell(row=row, column=6, value=book.copies_available)
            ws.cell(row=row, column=7, value=book.status)
            ws.cell(row=row, column=8, value=float(book.rating))
            ws.cell(row=row, column=9, value=book.language)
        
        # Adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_overdue_report():
        """Generate detailed overdue books report"""
        overdue_records = BorrowRecord.objects.filter(
            status='Overdue'
        ).select_related('user', 'book').order_by('due_date')
        
        return [{
            'user': record.user.username,
            'email': record.user.email,
            'book': record.book.title,
            'borrow_date': record.borrow_date,
            'due_date': record.due_date,
            'days_overdue': (timezone.now().date() - record.due_date).days,
            'fine': record.fine_amount,
            'fine_paid': record.fine_paid,
        } for record in overdue_records]

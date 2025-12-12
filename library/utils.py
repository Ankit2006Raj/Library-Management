"""Utility functions for exports, notifications, and more"""
from django.http import HttpResponse
from django.template.loader import render_to_string
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import io


def export_books_to_excel(books):
    """Export books to Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['ID', 'Title', 'Author', 'Category', 'ISBN', 'Publisher', 
               'Year', 'Language', 'Status', 'Available', 'Total', 'Rating']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for book in books:
        ws.append([
            book.id,
            book.title,
            book.author,
            book.category,
            book.isbn or '',
            book.publisher or '',
            book.published_year,
            book.language,
            book.status,
            book.copies_available,
            book.total_copies,
            float(book.rating)
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=books_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


def export_books_to_pdf(books):
    """Export books to PDF file"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph("Library Books Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Date
    date_text = Paragraph(
        f"Generated on: {datetime.now().strftime('%B %d, %Y')}",
        styles['Normal']
    )
    elements.append(date_text)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [['Title', 'Author', 'Category', 'Status', 'Rating']]
    for book in books:
        data.append([
            book.title[:30],
            book.author[:20],
            book.category,
            book.status,
            f"{book.rating}★"
        ])
    
    # Create table
    table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=books_{datetime.now().strftime("%Y%m%d")}.pdf'
    response.write(buffer.getvalue())
    return response


def export_borrow_history_to_excel(borrow_records):
    """Export borrow history to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borrow History"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['ID', 'User', 'Book', 'Borrow Date', 'Due Date', 'Return Date', 'Status', 'Fine']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    for record in borrow_records:
        ws.append([
            record.id,
            record.user.username,
            record.book.title,
            record.borrow_date.strftime('%Y-%m-%d'),
            record.due_date.strftime('%Y-%m-%d'),
            record.return_date.strftime('%Y-%m-%d') if record.return_date else 'Not Returned',
            record.status,
            float(record.fine_amount)
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=borrow_history_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


def send_due_date_reminders():
    """Send reminders for books due soon (to be called by Celery task)"""
    from django.utils import timezone
    from datetime import timedelta
    from .models import BorrowRecord, Notification
    
    tomorrow = timezone.now().date() + timedelta(days=1)
    
    # Get books due tomorrow
    due_soon = BorrowRecord.objects.filter(
        due_date=tomorrow,
        status='Borrowed'
    )
    
    for record in due_soon:
        Notification.objects.create(
            user=record.user,
            notification_type='DUE_SOON',
            title='Book Due Tomorrow',
            message=f'The book "{record.book.title}" is due tomorrow!',
            link=f'/my-books/'
        )


def send_overdue_notifications():
    """Send notifications for overdue books (to be called by Celery task)"""
    from django.utils import timezone
    from .models import BorrowRecord, Notification
    
    overdue_records = BorrowRecord.objects.filter(
        status='Overdue'
    )
    
    for record in overdue_records:
        # Check if notification already sent today
        today = timezone.now().date()
        existing = Notification.objects.filter(
            user=record.user,
            notification_type='OVERDUE',
            created_at__date=today,
            message__contains=record.book.title
        ).exists()
        
        if not existing:
            fine = record.calculate_fine()
            Notification.objects.create(
                user=record.user,
                notification_type='OVERDUE',
                title='Overdue Book',
                message=f'The book "{record.book.title}" is overdue. Fine: ${fine}',
                link=f'/my-books/'
            )


def calculate_library_statistics():
    """Calculate comprehensive library statistics"""
    from .models import Book, BorrowRecord, Review, User
    from django.db.models import Count, Avg, Sum
    
    stats = {
        'total_books': Book.objects.count(),
        'total_users': User.objects.count(),
        'total_borrows': BorrowRecord.objects.count(),
        'active_borrows': BorrowRecord.objects.filter(status='Borrowed').count(),
        'overdue_books': BorrowRecord.objects.filter(status='Overdue').count(),
        'total_reviews': Review.objects.count(),
        'average_rating': Review.objects.aggregate(Avg('rating'))['rating__avg'] or 0,
        'most_borrowed_books': BorrowRecord.objects.values('book__title').annotate(
            count=Count('id')
        ).order_by('-count')[:5],
        'most_active_users': BorrowRecord.objects.values('user__username').annotate(
            count=Count('id')
        ).order_by('-count')[:5],
        'category_distribution': Book.objects.values('category').annotate(
            count=Count('id')
        ).order_by('-count'),
    }
    
    return stats
